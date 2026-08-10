"""PDF/XLSX renderer for the Orbital Risk Factor Sheet — single A4 page.

Takes the assessment dict from services.orbital_risk and renders a document an
underwriter can drop into a file: identification, the headline exposure figure,
the components behind it, and — deliberately prominent — the scope boundaries.

Same reportlab pattern as services/report_pdf.py (no new dependencies).
"""
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

NAVY = colors.HexColor("#0A2540")
AMBER = colors.HexColor("#E0902E")
AMBER_BG = colors.HexColor("#FDF6EC")
AMBER_BD = colors.HexColor("#E8C48A")
SUB = colors.HexColor("#5A6B7B")
RULE = colors.HexColor("#D6E0EA")
INK = colors.HexColor("#1A2733")
BOX_BG = colors.HexColor("#F4F7FB")

S_TITLE = ParagraphStyle("t", fontName="Helvetica-Bold", fontSize=16, textColor=NAVY, leading=20)
S_KICK = ParagraphStyle("k", fontName="Helvetica-Bold", fontSize=7.5, textColor=AMBER, leading=10)
S_SUB = ParagraphStyle("s", fontName="Helvetica", fontSize=8.5, textColor=SUB, leading=11)
S_META = ParagraphStyle("m", fontName="Courier", fontSize=7.5, textColor=SUB, leading=11, alignment=2)
S_H = ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=9.5, textColor=NAVY,
                     leading=12, spaceBefore=9, spaceAfter=3)
S_BODY = ParagraphStyle("b", fontName="Helvetica", fontSize=8.5, textColor=INK, leading=11)
S_NOTE = ParagraphStyle("n", fontName="Helvetica", fontSize=7.5, textColor=SUB, leading=10)
S_BIG = ParagraphStyle("big", fontName="Courier-Bold", fontSize=24, textColor=AMBER,
                       leading=27, alignment=1)
S_BIGL = ParagraphStyle("bigl", fontName="Helvetica", fontSize=7.5, textColor=SUB,
                        leading=10, alignment=1)
S_BOXN = ParagraphStyle("bn", fontName="Courier-Bold", fontSize=13, textColor=NAVY,
                        leading=16, alignment=1)
S_BOXL = ParagraphStyle("bl", fontName="Helvetica", fontSize=6.5, textColor=SUB,
                        leading=9, alignment=1)
S_CELL = ParagraphStyle("c", fontName="Helvetica", fontSize=8, textColor=INK, leading=10)
S_CELLS = ParagraphStyle("cs", fontName="Helvetica", fontSize=6.8, textColor=SUB, leading=8.5)


def _sci(x: Optional[float]) -> str:
    if x is None:
        return "—"
    try:
        return f"{x:.3e}"
    except Exception:
        return "—"


def _pct(x: Optional[float]) -> str:
    if x is None:
        return "—"
    return f"{x:+.1f}%/yr"


def _report_id(a: Dict[str, Any]) -> str:
    """Deterministic short id from orbit + date (same input -> same id)."""
    import hashlib
    o = a.get("orbit", {})
    seed = f"{o.get('altitude_km')}|{o.get('inclination_deg')}|{datetime.now(timezone.utc):%Y%m%d}"
    return "#" + hashlib.sha256(seed.encode()).hexdigest()[:4].upper()


def render_insurance_pdf(a: Dict[str, Any], *, org: Optional[str] = None,
                         demo: bool = False, report_id: Optional[str] = None) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=16 * mm, rightMargin=16 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm,
                            title="Orbital Risk Factor Sheet")
    o = a.get("orbit", {})
    c = a.get("catalogue", {})
    b = a.get("burden", {})
    cas = a.get("cascade", {})
    tr = (a.get("trend") or {}).get("primary") or {}
    trb = (a.get("trend") or {}).get("band_based") or {}
    fx = a.get("debris_flux", {})
    bnd = a.get("boundaries", {})
    pctl = a.get("percentile", {}) or {}
    subj = a.get("subject", {}) or {}
    rid = report_id or _report_id(a)
    now = datetime.now(timezone.utc).strftime("%d %b %Y · %H:%M UTC")

    inc = o.get("inclination_deg")
    inc_s = f"{inc:.1f}°" if inc is not None else "all inclinations"
    gated = "inclination-gated" if c.get("inclination_gated") else "band-based"

    F: List[Any] = []

    # ── header ──
    head = Table([[
        Paragraph("CAS PLATFORM<br/><font size=6.5 color='#5A6B7B'>ORBITAL RISK INTELLIGENCE</font>",
                  ParagraphStyle("hh", fontName="Helvetica-Bold", fontSize=9,
                                 textColor=NAVY, leading=12)),
        Paragraph(f"REPORT <b>{rid}</b><br/>{now}<br/>{gated}", S_META),
    ]], colWidths=[95 * mm, 83 * mm])
    head.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                              ("LEFTPADDING", (0, 0), (-1, -1), 0),
                              ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))
    F.append(head)
    F.append(Spacer(1, 3 * mm))
    F.append(HRFlowable(width="100%", thickness=1.2, color=AMBER, spaceAfter=5 * mm))

    F.append(Paragraph("ORBITAL RISK FACTOR SHEET", S_KICK))
    if subj.get("name"):
        F.append(Paragraph(subj["name"], S_TITLE))
        bits = [f"NORAD {subj['norad']}"] if subj.get("norad") else []
        if subj.get("country"):
            bits.append(subj["country"])
        if subj.get("mass_kg"):
            bits.append(f"{subj['mass_kg']:,.0f} kg")
        if subj.get("rcs_size"):
            bits.append(f"{subj['rcs_size'].title()} radar cross-section")
        if subj.get("launch_date"):
            bits.append(f"launched {subj['launch_date'][:4]}")
        bits.append(f"{o.get('altitude_km')} km / {inc_s}")
        F.append(Paragraph(" &middot; ".join(bits), S_SUB))
    else:
        F.append(Paragraph(f"{o.get('altitude_km')} km / {inc_s}", S_TITLE))
    if org:
        F.append(Paragraph(f"Prepared for {org}", S_SUB))
    F.append(Spacer(1, 5 * mm))

    # ── headline ──
    F.append(Paragraph(_sci(b.get("lambda_per_year")), S_BIG))
    F.append(Paragraph("KINETIC BURDEN · expected collisions per year "
                       "per 10 m² reference cross-section", S_BIGL))
    if pctl.get("available"):
        pv = pctl.get("percentile", 0)
        pv_txt = "less than 1" if pv < 1 else str(pv)
        bar_w = 150 * mm
        fill = Table([[""]], colWidths=[bar_w * max(pv, 1) / 100.0],
                     rowHeights=[3.2 * mm])
        fill.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), AMBER),
                                  ("LEFTPADDING", (0, 0), (-1, -1), 0),
                                  ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                                  ("TOPPADDING", (0, 0), (-1, -1), 0),
                                  ("BOTTOMPADDING", (0, 0), (-1, -1), 0)]))
        track = Table([[fill]], colWidths=[bar_w], hAlign="CENTER")
        track.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EEF2F7")),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("BOX", (0, 0), (-1, -1), 0.4, RULE)]))
        F.append(Spacer(1, 3.5 * mm))
        F.append(track)
        F.append(Spacer(1, 1.5 * mm))
        F.append(Paragraph(
            "Higher collision burden than <b>" + pv_txt + "%</b> of sampled LEO shells "
            "(" + str(pctl.get("sample_shells", 0)) + " shells, 300\u20131400 km)",
            S_BIGL))
    F.append(Spacer(1, 5 * mm))

    # ── three boxes ──
    clr_y = cas.get("cloud_clearing_years_90pct")
    boxes = Table([[
        Paragraph(f"{c.get('threat', '—')}", S_BOXN),
        Paragraph(_pct(tr.get("cagr_pct_per_year")) if tr.get("available") else "—", S_BOXN),
        Paragraph(f"{clr_y:.0f} yr" if clr_y else "—", S_BOXN),
    ], [
        Paragraph(f"THREAT OBJECTS<br/>debris {c.get('debris', 0)} · R/B {c.get('rocket_body', 0)}", S_BOXL),
        Paragraph(f"THREAT TREND<br/>{tr.get('first_year', '')}–{tr.get('last_year', '')} "
                  f"· band {_pct(trb.get('cagr_pct_per_year')) if trb.get('available') else '—'}", S_BOXL),
        Paragraph(f"CLOUD PERSISTENCE<br/>{cas.get('exposed_pool', 0)} objects exposed", S_BOXL),
    ]], colWidths=[59 * mm, 59 * mm, 60 * mm])
    boxes.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BOX_BG),
        ("GRID", (0, 0), (-1, -1), 3, colors.white),
        ("TOPPADDING", (0, 0), (-1, 0), 7), ("BOTTOMPADDING", (0, 0), (-1, 0), 1),
        ("TOPPADDING", (0, 1), (-1, 1), 0), ("BOTTOMPADDING", (0, 1), (-1, 1), 7),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    F.append(boxes)

    # ── components ──
    F.append(Paragraph("COMPONENTS", S_H))
    rows = [["Component", "Value", "Basis"]]
    rows.append([Paragraph("Threat density (ρ)", S_CELL), _sci(b.get("rho_threat_per_km3")),
                 Paragraph("Non-manoeuvrable objects per km³ (debris + rocket bodies)", S_CELLS)])
    rows.append([Paragraph("Mean relative velocity", S_CELL),
                 f"{b.get('v_rel_mean_km_s', 0):.2f} km/s",
                 Paragraph("Isotropic encounter approximation", S_CELLS)])
    rows.append([Paragraph("Threat fraction", S_CELL), f"{c.get('threat_fraction_pct', 0)}%",
                 Paragraph("Share of the shell population that cannot manoeuvre", S_CELLS)])
    rows.append([Paragraph("Shell population", S_CELL), f"{c.get('total', 0)}",
                 Paragraph(f"payload {c.get('payload', 0)} · debris {c.get('debris', 0)} "
                           f"· R/B {c.get('rocket_body', 0)}", S_CELLS)])
    _fm = cas.get("fragmenting_mass_kg")
    _mb = cas.get("mass_basis") or "model default"
    rows.append([Paragraph("Cascade exposure", S_CELL), f"{cas.get('exposed_pool', 0)}",
                 Paragraph("Objects within ±50 km · NASA Standard Breakup Model + drag", S_CELLS)])
    rows.append([Paragraph("Fragmenting mass", S_CELL),
                 (f"{_fm:,.0f} kg" if _fm else "—"),
                 Paragraph(f"Basis: {_mb}" + (
                     f" · {cas.get('band_mass_profile',{}).get('objects_with_mass',0)} of "
                     f"{cas.get('band_mass_profile',{}).get('objects_in_band',0)} band objects "
                     "have a published mass (ESA DISCOS)"
                     if cas.get("band_mass_profile") else ""), S_CELLS)])
    if tr.get("available"):
        rows.append([Paragraph("Environment trend", S_CELL), _pct(tr.get("cagr_pct_per_year")),
                     Paragraph(f"{tr.get('threat_first')} → {tr.get('threat_last')} threat objects "
                               f"({tr.get('first_year')}–{tr.get('last_year')}), {tr.get('mode')}", S_CELLS)])
    else:
        rows.append([Paragraph("Environment trend", S_CELL), "n/a",
                     Paragraph(str(tr.get("note", "not available"))[:110], S_CELLS)])
    rows.append([Paragraph("Lethal non-trackable flux", S_CELL), "not included",
                 Paragraph("1–10 cm debris requires ESA MASTER / NASA ORDEM — see scope", S_CELLS)])

    t = Table(rows, colWidths=[44 * mm, 28 * mm, 106 * mm], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7.5),
        ("FONTNAME", (1, 1), (1, -1), "Courier"),
        ("FONTSIZE", (1, 1), (1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, RULE),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
    ]))
    F.append(t)

    # ── scope / boundaries ──
    F.append(Paragraph("SCOPE OF THIS ASSESSMENT", S_H))
    scope = Table([[Paragraph(
        f"<b>What this is.</b> {bnd.get('is', '')}<br/><br/>"
        f"<b>What this is not.</b> {bnd.get('is_not', '')}<br/><br/>"
        f"<b>Coverage.</b> {bnd.get('coverage', '')}", S_BODY)]],
        colWidths=[178 * mm])
    scope.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), AMBER_BG),
        ("BOX", (0, 0), (-1, -1), 0.8, AMBER_BD),
        ("LINEBEFORE", (0, 0), (0, -1), 2.5, AMBER),
        ("TOPPADDING", (0, 0), (-1, -1), 7), ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    F.append(scope)

    F.append(Spacer(1, 4 * mm))
    F.append(HRFlowable(width="100%", thickness=0.5, color=RULE, spaceAfter=2.5 * mm))
    F.append(Paragraph(
        "Data sources: Space-Track public catalogue (current state) and the CAS catalogue archive "
        "(2020–2026 trend). Methodology is documented in the CAS Methodology Note and is reproducible: "
        "the same inputs yield the same figures. Generated by CAS Platform · casplatform.com",
        S_NOTE))

    def _watermark(canvas, _doc):
        canvas.saveState()
        if demo:
            canvas.setFont("Helvetica-Bold", 62)
            canvas.setFillColor(colors.Color(0.88, 0.56, 0.18, alpha=0.13))
            canvas.translate(A4[0] / 2, A4[1] / 2)
            canvas.rotate(32)
            canvas.drawCentredString(0, 0, "DEMO — SYNTHETIC DATA")
            canvas.rotate(-32)
            canvas.translate(-A4[0] / 2, -A4[1] / 2)
        canvas.setFont("Helvetica", 6.5)
        canvas.setFillColor(SUB)
        canvas.drawString(16 * mm, 8 * mm, f"CAS PLATFORM · ORBITAL RISK FACTOR SHEET · {rid}")
        canvas.drawRightString(A4[0] - 16 * mm, 8 * mm,
                               "DEMO ACCOUNT — NOT FOR UNDERWRITING USE" if demo else "Page 1 of 1")
        canvas.restoreState()

    doc.build(F, onFirstPage=_watermark, onLaterPages=_watermark)
    return buf.getvalue()


def render_insurance_xlsx(a: Dict[str, Any], *, org: Optional[str] = None,
                          demo: bool = False, report_id: Optional[str] = None) -> bytes:
    """Same figures as a spreadsheet, for the underwriter's own model."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    o = a.get("orbit", {}); c = a.get("catalogue", {}); b = a.get("burden", {})
    cas = a.get("cascade", {}); bnd = a.get("boundaries", {})
    tr = (a.get("trend") or {}).get("primary") or {}
    trb = (a.get("trend") or {}).get("band_based") or {}
    rid = report_id or _report_id(a)

    wb = Workbook(); ws = wb.active; ws.title = "Risk Factor Sheet"
    hdr = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="0A2540")

    def row(k, v, note=""):
        ws.append([k, v, note])

    ws.append(["ORBITAL RISK FACTOR SHEET", "", ""])
    ws["A1"].font = Font(bold=True, size=14, color="0A2540")
    row("Report", rid, datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))
    if org: row("Prepared for", org, "")
    if demo: row("NOTICE", "DEMO — SYNTHETIC DATA", "Not for underwriting use")
    row("", "", "")
    ws.append(["PARAMETER", "VALUE", "BASIS"])
    for cell in ws[ws.max_row]:
        cell.font = hdr; cell.fill = fill
    row("Altitude (km)", o.get("altitude_km"), "")
    row("Inclination (deg)", o.get("inclination_deg") if o.get("inclination_deg") is not None else "all", "")
    row("Mode", "inclination-gated" if c.get("inclination_gated") else "band-based", "")
    row("Kinetic burden (/yr)", b.get("lambda_per_year"), "collisions/yr per 10 m² reference")
    _p = a.get("percentile", {}) or {}
    row("LEO percentile", _p.get("percentile") if _p.get("available") else "n/a",
        "vs " + str(_p.get("sample_shells", "?")) + " sampled shells, 300-1400 km")
    row("Threat objects", c.get("threat"), "debris + rocket bodies (non-manoeuvrable)")
    row("Debris", c.get("debris"), "")
    row("Rocket bodies", c.get("rocket_body"), "")
    row("Payloads", c.get("payload"), "")
    row("Shell total", c.get("total"), "")
    row("Threat fraction (%)", c.get("threat_fraction_pct"), "")
    row("Threat density (/km3)", b.get("rho_threat_per_km3"), "")
    row("Mean rel. velocity (km/s)", b.get("v_rel_mean_km_s"), "isotropic approximation")
    row("Trend (%/yr)", tr.get("cagr_pct_per_year") if tr.get("available") else "n/a",
        f"{tr.get('threat_first', '')} -> {tr.get('threat_last', '')} ({tr.get('mode', '')})")
    row("Trend band-based (%/yr)", trb.get("cagr_pct_per_year") if trb.get("available") else "n/a", "")
    row("Cascade exposed pool", cas.get("exposed_pool"), "objects within ±50 km")
    row("Cloud persistence (yr)", cas.get("cloud_clearing_years_90pct"), "90% clearing, NASA SBM + drag")
    row("Lethal non-trackable flux", "not included", "requires ESA MASTER / NASA ORDEM")
    row("", "", "")
    ws.append(["SCOPE", "", ""])
    for cell in ws[ws.max_row]:
        cell.font = hdr; cell.fill = fill
    for label, key in [("What this is", "is"), ("What this is not", "is_not"), ("Coverage", "coverage")]:
        ws.append([label, bnd.get(key, ""), ""])
        ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 52
    out = BytesIO(); wb.save(out)
    return out.getvalue()
