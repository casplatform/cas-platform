"""XLSX renderer for monthly/annual reports — openpyxl, multi-sheet.

Sheets: Summary, Satellites (or Top Objects), Decisions, Space Weather,
and Monthly Trend (annual only). Mirrors the JSON report from
services.reporting so operators can analyse the data in a spreadsheet.
"""
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

NAVY = "0A2540"
BLUE = "1B6CA8"
CYAN = "15A0C8"
AMBER_BG = "FFF7E8"
LIGHT = "F2F7FB"

_HDR_FILL = PatternFill("solid", fgColor=BLUE)
_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
_TITLE_FONT = Font(bold=True, color=NAVY, size=14)
_SUB_FONT = Font(color="5A6B7B", size=9)
_LABEL_FONT = Font(bold=True, color=NAVY, size=10)
_THIN = Side(style="thin", color="D6E0EA")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")


def _pc(v: Optional[float]) -> str:
    return f"{v:.2e}" if v is not None else "—"


def _num(v):
    return v if v is not None else "—"


def _norad(v):
    """NORAD id'yi Excel'de sayi olarak yaz; bos/gecersizse string birak."""
    try:
        return int(str(v).strip())
    except (TypeError, ValueError):
        return v if v is not None else "—"


def _style_header(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.border = _BORDER


def _table(ws, start_row: int, headers: List[str], rows: List[List[Any]]) -> int:
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    _style_header(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for j, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.border = _BORDER
        r += 1
    return r


def _autosize(ws, widths: Dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _header_block(ws, report: Dict[str, Any], title: str):
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    prov = report.get("provenance", {})
    ow = prov.get("observation_window", {}) or {}
    ws["A2"] = (f"Scope: {report['scope']['mode']} "
                f"({report['scope'].get('satellites', 0)} satellites)  ·  "
                f"Generated: {str(prov.get('generated_at',''))[:16]}Z")
    ws["A2"].font = _SUB_FONT
    ws["A3"] = (f"Observation window: {str(ow.get('first_observation',''))[:10]} → "
                f"{str(ow.get('last_observation',''))[:10]} "
                f"({ow.get('days_observing','—')} days)  ·  Period: {report['period']['label']}")
    ws["A3"].font = _SUB_FONT


def render_report_xlsx(report: Dict[str, Any]) -> bytes:
    wb = Workbook()
    rtype = report["report_type"].title()
    label = report["period"]["label"]

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    _header_block(ws, report, f"CAS {rtype} Report — {label}")
    s = report.get("summary", {})
    r = _table(ws, 5,
               ["Metric", "Value"],
               [["Unique conjunctions", _num(s.get("unique_conjunctions"))],
                ["CDM updates", _num(s.get("cdm_updates"))],
                ["Max Pc", _pc(s.get("max_pc"))],
                ["Min miss (m)", _num(s.get("min_miss_m"))],
                ["Median miss (m)", _num(s.get("median_miss_m"))]])
    # risk breakdown
    r += 1
    ws.cell(row=r, column=1, value="Risk tier").font = _LABEL_FONT
    r += 1
    by_risk = s.get("by_risk", {}) or {}
    r = _table(ws, r, ["Risk", "Count"], [[k, v] for k, v in by_risk.items()] or [["—", 0]])
    # decisions summary
    r += 1
    d = report.get("decisions", {})
    ws.cell(row=r, column=1, value="Decision activity").font = _LABEL_FONT
    r += 1
    ba = d.get("by_action", {}) or {}
    r = _table(ws, r, ["Recorded action", "Count"],
               [[k, v] for k, v in ba.items()] or [["No actions recorded", 0]])
    mrh = d.get("median_response_hours")
    ws.cell(row=r, column=1, value="Median response (h)")
    ws.cell(row=r, column=2, value=(round(mrh, 1) if mrh is not None else "insufficient sample"))
    # space weather
    r += 2
    ws.cell(row=r, column=1, value="Space weather").font = _LABEL_FONT
    r += 1
    w = report.get("space_weather", {})
    r = _table(ws, r, ["Metric", "Value"],
               [["Snapshots", _num(w.get("snapshots"))],
                ["Kp avg", _num(w.get("kp_avg"))],
                ["Kp max", _num(w.get("kp_max"))],
                ["F10.7 max", _num(w.get("f107_max"))],
                ["Days Kp ≥ 5", _num(w.get("elevated_days_kp_ge5"))]])
    _autosize(ws, {1: 34, 2: 26})

    # ── Sheet 2: Satellites (operator) or Top Objects (admin) ──
    sats = report.get("satellites")
    if sats is not None:
        ws2 = wb.create_sheet("Satellites")
        ws2["A1"] = "Your Satellites — Activity Breakdown"
        ws2["A1"].font = _TITLE_FONT
        _table(ws2, 3, ["Satellite", "NORAD", "Conjunctions", "Max Pc", "Min miss (m)", "Last TCA"],
               [[x["sat_name"], _norad(x["norad_id"]), x["unique_conjunctions"],
                 _pc(x["max_pc"]), _num(x["min_miss_m"]), str(x.get("last_tca", ""))[:16]]
                for x in sats] or [["No activity", "", "", "", "", ""]])
        _autosize(ws2, {1: 34, 2: 12, 3: 14, 4: 12, 5: 14, 6: 18})

        tc = report.get("top_counterparties") or []
        ws3 = wb.create_sheet("Counterparties")
        ws3["A1"] = "Top Counterparty Objects"
        ws3["A1"].font = _TITLE_FONT
        _table(ws3, 3, ["Object", "NORAD", "Conjunctions", "Max Pc"],
               [[x["name"], _norad(x["norad_id"]), x["unique_conjunctions"], _pc(x["max_pc"])]
                for x in tc] or [["—", "", "", ""]])
        _autosize(ws3, {1: 40, 2: 12, 3: 14, 4: 12})
    else:
        to = report.get("top_objects") or []
        ws2 = wb.create_sheet("Top Objects")
        ws2["A1"] = "Most Conjunction-Active Objects (global)"
        ws2["A1"].font = _TITLE_FONT
        _table(ws2, 3, ["Object", "NORAD", "Conjunctions", "Max Pc"],
               [[x["name"], _norad(x["norad_id"]), x["unique_conjunctions"], _pc(x["max_pc"])]
                for x in to] or [["—", "", "", ""]])
        _autosize(ws2, {1: 40, 2: 12, 3: 14, 4: 12})

    # ── Sheet 3: Monthly Trend (annual only) ──
    trend = report.get("monthly_trend")
    if trend:
        ws4 = wb.create_sheet("Monthly Trend")
        ws4["A1"] = f"Month-by-Month Trend — {label}"
        ws4["A1"].font = _TITLE_FONT
        _table(ws4, 3, ["Month", "Unique conjunctions", "Max Pc", "Min miss (m)", "Actions"],
               [[f"{t['month']:02d}", t["unique_conjunctions"], _pc(t["max_pc"]),
                 _num(t["min_miss_m"]), t["actions"]] for t in trend])
        _autosize(ws4, {1: 8, 2: 20, 3: 14, 4: 14, 5: 10})

    # ── Sheet: Notes ──
    wsn = wb.create_sheet("Notes")
    wsn["A1"] = "Scope & Honesty Notes"
    wsn["A1"].font = _TITLE_FONT
    for i, note in enumerate(report.get("notes", []), start=3):
        c = wsn.cell(row=i, column=1, value="• " + note)
        c.alignment = _WRAP
    wsn.column_dimensions["A"].width = 110

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
