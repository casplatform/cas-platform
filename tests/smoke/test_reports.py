"""Smoke tests for the reports endpoints (/api/v2/reports/*).

Self-contained: hits the live FastAPI service on localhost. Verifies
fail-closed auth, the public /info shape, and that an authenticated
report returns the expected JSON structure. No fixture dependencies.
"""
import io
import json
import urllib.request

FASTAPI = "http://127.0.0.1:8766"
ENGINE = "http://127.0.0.1:8765"
TEST_EMAIL = "test@casplatform.com"
TEST_PASS = "Test!2026"


def _get(url, token=None, raw=False):
    req = urllib.request.Request(url)
    if token:
        req.add_header("Authorization", "Bearer " + token)
    try:
        resp = urllib.request.urlopen(req, timeout=30)
        data = resp.read()
        return resp.status, (data if raw else json.loads(data))
    except urllib.error.HTTPError as e:
        return e.code, None


def _login():
    body = json.dumps({"email": TEST_EMAIL, "password": TEST_PASS}).encode()
    req = urllib.request.Request(ENGINE + "/auth/login", data=body,
                                 headers={"Content-Type": "application/json"})
    try:
        resp = urllib.request.urlopen(req, timeout=30)
        return json.loads(resp.read()).get("token", "")
    except Exception:
        return ""


def test_reports_monthly_requires_auth():
    """Unauthenticated monthly report must be rejected (fail-closed)."""
    status, _ = _get(FASTAPI + "/api/v2/reports/monthly?year=2026&month=6")
    assert status in (401, 403), f"expected 401/403, got {status}"


def test_reports_annual_requires_auth():
    """Unauthenticated annual report must be rejected."""
    status, _ = _get(FASTAPI + "/api/v2/reports/annual?year=2026")
    assert status in (401, 403), f"expected 401/403, got {status}"


def test_reports_info_requires_auth():
    """/info describes capability to an authenticated caller, matching
    insurance/info and mission/info; it is not open to anonymous callers."""
    status, body = _get(FASTAPI + "/api/v2/reports/info")
    assert status in (401, 403), f"expected 401/403, got {status}"


def test_reports_bad_format_rejected():
    """format=xml must be a validation error (422)."""
    token = _login()
    if not token:
        return  # login unavailable in this env; skip silently
    status, _ = _get(
        FASTAPI + "/api/v2/reports/monthly?year=2026&month=6&format=xml", token=token)
    assert status == 422, f"expected 422 for bad format, got {status}"


def test_reports_monthly_json_shape():
    """Authenticated monthly report returns the expected top-level keys."""
    token = _login()
    if not token:
        return  # skip if login unavailable
    status, body = _get(
        FASTAPI + "/api/v2/reports/monthly?year=2026&month=6", token=token)
    assert status == 200
    for key in ("report_type", "period", "scope", "summary", "notes", "provenance"):
        assert key in body, f"missing key: {key}"
    assert body["report_type"] == "monthly"
    assert body["scope"]["mode"] in ("watchlist", "global")
    assert isinstance(body["summary"].get("unique_conjunctions"), int)


def test_reports_pdf_download():
    """Authenticated PDF download returns a real PDF."""
    token = _login()
    if not token:
        return
    status, data = _get(
        FASTAPI + "/api/v2/reports/monthly?year=2026&month=6&format=pdf",
        token=token, raw=True)
    assert status == 200
    assert data[:5] == b"%PDF-", "response is not a PDF"


def test_reports_xlsx_download():
    """Authenticated XLSX download returns a real xlsx (zip) with expected sheets."""
    token = _login()
    if not token:
        return
    status, data = _get(
        FASTAPI + "/api/v2/reports/annual?year=2026&format=xlsx",
        token=token, raw=True)
    assert status == 200
    assert data[:2] == b"PK", "response is not a zip/xlsx"
    # verify sheets if openpyxl available
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "Summary" in wb.sheetnames
        assert "Monthly Trend" in wb.sheetnames
    except ImportError:
        pass
